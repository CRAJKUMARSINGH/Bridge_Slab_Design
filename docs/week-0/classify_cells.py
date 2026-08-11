#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
classify_cells.py -- classify every interesting cell in a bridge-design Excel
workbook as one of:

    VARIABLE     an independent design input the user changes per run
                 (hard-coded number, has a label, feeds formulas, has NO formula
                  precedents of its own -> a ROOT of the dependency graph)
    COEFFICIENT  a fixed factor / material property / code limit / numeric
                 literal embedded inside a formula (constants of the model)
    CONSTRAINT   a design check / inequality the result must satisfy
                 (utilisation <= 1.0, deflection <= limit, IF(...<=...,"OK",...))
    INTERMEDIATE a calculated value that only feeds later formulas
    AMBIGUOUS    low-confidence item -> the script ASKS YOU to confirm

Two passes over the workbook:
    pass 1  openpyxl(data_only=False)   -> formulas, labels, defined names
    pass 2  openpyxl(data_only=True)    -> cached values written by Excel

Usage
-----
  # read-only report
  python classify_cells.py workbook.xlsx

  # ask you to confirm every ambiguous cell (type v/c/k/s per prompt)
  python classify_cells.py workbook.xlsx --confirm

  # batch mode: apply previously-supplied answers, never prompt
  python classify_cells.py workbook.xlsx --answers my_answers.json

  # write JSON outputs to ./out
  python classify_cells.py workbook.xlsx --confirm --outdir ./out

answers.json format:
  {"overrides": {"Calc!B23": "coefficient", "Inputs!B12": "variable"},
   "skip": ["Sheet1!A5"]}
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict

import openpyxl

# ---------------------------------------------------------------------------
# regexes & keyword sets
# ---------------------------------------------------------------------------
REF_RE = re.compile(
    r"(?<![\w])(?:'?([^'!]+)'?!)?(\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)"
)
NUM_IN_FORMULA = re.compile(r"(?<![\w.])(\d+(?:\.\d+)?(?:[Ee][+-]?\d+)?)")

VARIABLE_SHEET_WORDS = ("input", "param", "variable")
COEFF_SHEET_WORDS = ("factor", "coeff", "constant", "limit", "code", "table", "material")
CHOSEN_WORDS = ("chosen", "selected", "adopted", "provided")          # engineer decision -> variable
COEFF_WORDS = ("factor", "gamma", "modulus", "density", "unit weight",
               "limit", "coefficient", "constant", "partial safety", "capacity")
VARIABLE_WORDS = ("span", "length", "width", "depth", "thickness", "spacing",
                  "number of", "strength", "grade", "load")
CHECK_WORDS = ("check", "util", "ok", "govern", "verif", "capacity", "limit state")

UNIT_TOKENS = ("kN/m", "kN*m", "kNm", "kN.m", "N/mm2", "mm4", "mm3", "mm2",
               "kN", "MPa", "kPa", "mm", "km", "m3", "m2", "%", "m")

# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------
def sheet_key(name: str) -> str:
    return re.sub(r"[ _\-]", "", name.lower())


def cell_to_rc(cell: str):
    m = re.match(r"\$?([A-Z]{1,3})\$?(\d+)$", cell)
    if not m:
        return None
    col = 0
    for ch in m.group(1):
        col = col * 26 + (ord(ch) - 64)
    return col, int(m.group(2))


def rng_to_cells(sheet: str, rng: str, cap: int = 10000) -> list:
    """'A1', '$A$1' or 'A1:B5' -> ['Sheet!A1', ...] (upper-case cells)"""
    rng = rng.replace("$", "")
    cells = []
    if ":" in rng:
        a, b = rng.split(":", 1)
        ca, ra = cell_to_rc(a)
        cb, rb = cell_to_rc(b)
        if not (ca and cb):
            return []
        for r in range(ra, rb + 1):
            for c in range(ca, cb + 1):
                cells.append(f"{sheet}!{openpyxl.utils.get_column_letter(c)}{r}")
                if len(cells) >= cap:
                    return cells
    else:
        cells.append(f"{sheet}!{rng.upper()}")
    return cells


UNIT_PAREN_RE = re.compile(r"\(([^()]*)\)")


def extract_units(label: str) -> str:
    if not label:
        return ""
    for grp in UNIT_PAREN_RE.findall(label):
        g = grp.strip()
        if g in ("-", ""):
            return "-"
        low = g
        for tok in UNIT_TOKENS:
            if tok in low:
                return tok
        if len(g) <= 8 and not re.search(r"[a-z]{4,}", g):
            return g
    return ""


def short_label(label: str, n: int = 28) -> str:
    if not label:
        return "(no label)"
    clean = re.sub(r"\s+", " ", label.strip())
    if " (" in clean:
        clean = clean.split(" (")[0]
    return clean[:n]


# ---------------------------------------------------------------------------
# main classifier
# ---------------------------------------------------------------------------
class Cell:
    __slots__ = ("id", "sheet", "cell", "label", "units", "formula", "value",
                 "is_formula", "refs", "refd_by", "precedent_ids",
                 "defined_names", "kind", "detail", "confidence", "status")

    def __init__(self, cid, sheet, cell):
        self.id = cid
        self.sheet = sheet
        self.cell = cell
        self.label = ""
        self.units = ""
        self.formula = None
        self.value = None
        self.is_formula = False
        self.refs = []
        self.refd_by = set()
        self.precedent_ids = []
        self.defined_names = []
        self.kind = None          # variable | coefficient | constraint | intermediate | ambiguous | orphan | noise
        self.detail = ""          # why
        self.confidence = 0.0
        self.status = "auto"      # auto | confirmed-by-user | skipped


def classify_workbook(path: str, answers: dict | None, min_conf: float, outdir: str):
    answers = answers or {}
    overrides = {k.upper(): v for k, v in answers.get("overrides", {}).items()}
    skip_ids = {k.upper() for k in answers.get("skip", [])}

    wb_f = openpyxl.load_workbook(path, data_only=False)
    wb_v = openpyxl.load_workbook(path, data_only=True)

    cells: dict[str, Cell] = {}
    label_seed: dict[str, str] = {}

    # ---- pass 1: structure + formulas ------------------------------------
    for ws in wb_f.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                cid = f"{ws.title}!{c.coordinate.upper()}"
                cell = Cell(cid, ws.title, c.coordinate.upper())
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    cell.is_formula = True
                    cell.formula = v[1:]
                else:
                    cell.value = v
                    if isinstance(v, str) and len(v) > 1 and re.search(r"[A-Za-z]", v):
                        label_seed[cid] = v          # text cells can serve as labels
                cells[cid] = cell

    # ---- pass 2: cached computed values -----------------------------------
    for ws in wb_v.worksheets:
        for row in ws.iter_rows():
            for c in row:
                cid = f"{ws.title}!{c.coordinate.upper()}"
                if cid in cells and c.value is not None:
                    cells[cid].value = c.value

    # ---- labels: left neighbour first, then cell above -------------------
    for cid, cell in list(cells.items()):
        if cell.label:
            continue
        ws, coord = cell.sheet, cell.cell
        col, row = cell_to_rc(coord)
        if not col:
            continue
        for dc in range(1, 7):                       # look left
            nid = f"{ws}!{openpyxl.utils.get_column_letter(max(1, col - dc))}{row}"
            if nid in label_seed:
                cell.label = label_seed[nid]
                break
        if not cell.label:
            for dr in range(1, 9):                   # look above
                nid = f"{ws}!{openpyxl.utils.get_column_letter(col)}{max(1, row - dr)}"
                if nid in label_seed:
                    cell.label = label_seed[nid]
                    break
        cell.units = extract_units(cell.label)

    # ---- defined names (typical for model inputs) -------------------------
    defined_map: dict[str, list] = {}
    try:
        for name in wb_f.defined_names:
            dn = wb_f.defined_names[name]
            attr = getattr(dn, "attr_text", None) or ""
            ids = []
            for m in REF_RE.finditer(attr):
                sh = (m.group(1) or "").strip("'") or ""
                ids += rng_to_cells(sh, m.group(2))
            if not ids:                              # LibreOffice syntax 'Inputs.$B.$3'
                norm = attr.replace("$", "").replace(".", "!")
                for m in REF_RE.finditer(norm):
                    sh = (m.group(1) or "").strip("'") or ""
                    ids += rng_to_cells(sh, m.group(2))
            defined_map[name.lower()] = ids
            for cid in ids:
                if cid in cells:
                    cells[cid].defined_names.append(name)
    except Exception as e:                           # pragma: no cover
        print(f"[warn] defined names unreadable: {e}")

    # ---- dependency graph -------------------------------------------------
    def_resolve = {name.lower(): ids for name, ids in defined_map.items()}
    for cid, cell in cells.items():
        if not cell.is_formula:
            continue
        refs = []
        for m in REF_RE.finditer(cell.formula):
            sh = (m.group(1) or "").strip("'") or cell.sheet
            refs += rng_to_cells(sh, m.group(2))
        # bare defined-name identifiers used in the formula
        for tok in re.findall(r"[A-Za-z_][A-Za-z0-9_.]*", cell.formula):
            if tok.lower() in def_resolve and tok not in ("IF", "MAX", "MIN", "ABS", "TRUE", "FALSE"):
                refs += def_resolve[tok.lower()]
        cell.refs = sorted(set(refs))
        for r in refs:
            if r in cells:
                cells[r].refd_by.add(cid)

    for cell in cells.values():
        cell.refd_by = set(cell.refd_by)
        cell.precedent_ids = cell.refs

    # ---- classification: constants ----------------------------------------
    for cid, cell in cells.items():
        if cid in skip_ids:
            cell.kind, cell.status = "skipped", "skipped"
            continue
        if cell.is_formula:
            continue                                  # formulas handled below
        if not isinstance(cell.value, (int, float)):
            cell.kind, cell.detail = "noise", "text / non-numeric constant"
            continue
        refd = len(cell.refd_by) > 0
        sk = sheet_key(cell.sheet)
        lab = (cell.label or "").lower()
        if not refd and not any(w in sk for w in VARIABLE_SHEET_WORDS):
            cell.kind, cell.detail, cell.confidence = "orphan", "unreferenced constant (no formula uses it)", 0.6
            continue

        if any(w in sk for w in VARIABLE_SHEET_WORDS):
            cell.kind, cell.detail, cell.confidence = "variable", "sheet indicates input/parameter", 0.95
        elif any(w in sk for w in COEFF_SHEET_WORDS):
            cell.kind, cell.detail, cell.confidence = "coefficient", "sheet indicates factor/limit/table", 0.9
        elif any(w in lab for w in CHOSEN_WORDS):
            cell.kind, cell.detail, cell.confidence = "variable", "engineer-chosen design decision", 0.85
        elif any(w in lab for w in COEFF_WORDS):
            cell.kind, cell.detail, cell.confidence = "coefficient", "factor / property / capacity limit", 0.82
        elif any(w in lab for w in VARIABLE_WORDS):
            cell.kind, cell.detail, cell.confidence = "variable", "labelled design input", 0.8
        else:
            cell.kind, cell.detail, cell.confidence = "ambiguous", "cannot tell variable vs coefficient", 0.5
        if cell.defined_names and cell.kind == "ambiguous":
            cell.kind, cell.detail, cell.confidence = "variable", "has a defined name (input-like)", 0.75

    # ---- classification: formulas -----------------------------------------
    for cid, cell in cells.items():
        if cid in skip_ids or not cell.is_formula:
            continue
        lab = (cell.label or "").lower()
        up = cell.formula.upper()
        has_cond = ("IF(" in up) or ("<" in up) or (">" in up)
        check_hit = any(w in lab for w in CHECK_WORDS)
        if has_cond or check_hit:
            cell.kind, cell.detail, cell.confidence = "constraint", "design check / inequality", 0.85
        elif len(cell.refd_by) > 0:
            cell.kind, cell.detail, cell.confidence = "intermediate", "calculated value feeding other cells", 0.9
        else:
            cell.kind, cell.detail, cell.confidence = "intermediate", "calculated leaf result", 0.85

    # ---- numeric literals inside formulas -> embedded coefficients --------
    literals: Counter = Counter()
    literal_usage: dict[str, list] = defaultdict(list)
    for cid, cell in cells.items():
        if not cell.is_formula or not cell.formula:
            continue
        formula = cell.formula
        for m in NUM_IN_FORMULA.finditer(formula):
            if m.start() > 0 and formula[m.start() - 1] == "^":
                continue              # pure exponent (^2, ^3) -- not a coefficient
            lit = m.group(1)
            try:
                num = float(lit)
            except ValueError:
                continue
            key = repr(int(num)) if num == int(num) else repr(num)
            literals[key] += 1
            literal_usage[key].append({"cell": cid, "label": short_label(cell.label), "formula": cell.formula})

    # ---- apply user overrides (answers file) ------------------------------
    for cid, kind in overrides.items():
        if cid in cells and kind in ("variable", "coefficient", "constraint", "intermediate", "skip"):
            cells[cid].kind = "skipped" if kind == "skip" else kind
            cells[cid].detail = "assigned by user"
            cells[cid].confidence = 1.0
            cells[cid].status = "confirmed-by-user"

    # ---- confirmation stage -----------------------------------------------
    ambiguous = [c for c in cells.values()
                 if c.kind == "ambiguous" and c.id not in skip_ids]
    if ambiguous:
        if args.confirm and sys.stdin.isatty():
            print("\n= CONFIRMATION REQUIRED =============================================")
            print("Enter: <Enter> keep guess | v variable | c coefficient | k constraint | s skip")
            for c in ambiguous:
                print(f"  {c.id:<14} {short_label(c.label, 34):<34} value={c.value!r}")
                ans = input("      keep [v/c/k/s]: ").strip().lower()
                if ans in ("v", "variable"):
                    c.kind, c.detail, c.confidence, c.status = "variable", "confirmed by user", 1.0, "confirmed-by-user"
                elif ans in ("c", "coefficient"):
                    c.kind, c.detail, c.confidence, c.status = "coefficient", "confirmed by user", 1.0, "confirmed-by-user"
                elif ans in ("k", "constraint"):
                    c.kind, c.detail, c.confidence, c.status = "constraint", "confirmed by user", 1.0, "confirmed-by-user"
                elif ans in ("s", "skip"):
                    c.kind, c.status = "skipped", "skipped"
                # Enter -> keep 'ambiguous'
        else:
            print(f"\n[info] {len(ambiguous)} ambiguous cell(s) -- re-run with --confirm to"
                  f" resolve interactively, or supply --answers JSON.")

    # ---- constraint inequality extraction ---------------------------------
    constraints = []
    for cid, cell in cells.items():
        if cell.kind != "constraint":
            continue
        cond = None
        raw_formula = cell.formula or ""
        m = re.search(r"IF\(\s*(.*?)\s*,", raw_formula, flags=re.I | re.S)
        if m:
            cond = m.group(1).strip()
        else:
            m2 = re.search(r"(.+?)\s*(<=|>=|<|>|=)\s*(.+)", raw_formula)
            if m2:
                cond = f"{m2.group(1)} {m2.group(2)} {m2.group(3)}"
        g_form = ""
        if "util" in (cell.label or "").lower():
            cond = f"{short_label(cell.label)} <= 1.0"
            if isinstance(cell.value, (int, float)):
                g_form = f"g = {cell.value:.4g} - 1.0 <= 0   (feasible={cell.value <= 1.0!s})"
        def _lbl(s):
            s = s.replace("$", "")
            if s in cells:
                return short_label(cells[s].label)
            return s
        raw_cond = cond or ""
        labeled_cond = re.sub(r"\$?[A-Z]{1,3}\$?\d+", lambda mm: _lbl(mm.group(0)), raw_cond)
        constraints.append({
            "id": cid, "label": cell.label, "kind": "if_check" if m else "pairwise",
            "inequality": labeled_cond, "inequality_raw": raw_cond, "g_form": g_form,
            "value": cell.value, "formula": raw_formula,
            "confidence": cell.confidence, "status": cell.status,
        })

    # ---- assemble outputs ---------------------------------------------------
    variables = []
    for c in cells.values():
        if c.kind == "variable":
            variables.append({
                "id": c.id, "sheet": c.sheet, "cell": c.cell, "label": c.label,
                "units": c.units, "value": c.value,
                "defined_name": c.defined_names or None,
                "referenced_by": sorted(c.refd_by), "confidence": c.confidence,
                "status": c.status,
            })
    coeff_cells = []
    for c in cells.values():
        if c.kind == "coefficient":
            coeff_cells.append({
                "id": c.id, "sheet": c.sheet, "cell": c.cell, "label": c.label,
                "units": c.units, "value": c.value,
                "referenced_by": sorted(c.refd_by), "confidence": c.confidence,
                "status": c.status,
            })
    embedded = [{"literal": k, "occurrences": v, "used_in": literal_usage[k][:5]}
                for k, v in literals.most_common()]

    os.makedirs(outdir, exist_ok=True)
    registry = {
        "workbook": os.path.basename(path),
        "cells": [{
            "id": c.id, "kind": c.kind, "detail": c.detail, "confidence": c.confidence,
            "label": c.label, "units": c.units, "value": c.value,
            "formula": c.formula, "precedents": c.precedent_ids,
            "referenced_by": sorted(c.refd_by)[:10],
            "refd_by_count": len(c.refd_by),
            "defined_name": c.defined_names or None, "status": c.status,
        } for c in sorted(cells.values(), key=lambda x: x.id)],
    }
    with open(os.path.join(outdir, "registry.json"), "w") as fh:
        json.dump(registry, fh, indent=2)
    with open(os.path.join(outdir, "variables.json"), "w") as fh:
        json.dump(variables, fh, indent=2)
    with open(os.path.join(outdir, "coefficients.json"), "w") as fh:
        json.dump({"factor_and_limit_cells": coeff_cells, "embedded_literals": embedded}, fh, indent=2)
    with open(os.path.join(outdir, "constraints.json"), "w") as fh:
        json.dump(constraints, fh, indent=2)

    # ---- console report ------------------------------------------------------
    def _val(c):
        return c.value if c.value is not None else "(not cached - open workbook in Excel once)"

    print("\n" + "=" * 74)
    print(f" WORKBOOK: {os.path.basename(path)}")
    print("=" * 74)
    print(f"\n V A R I A B L E S   (independent design inputs)   [{len(variables)}]\n")
    for i, v in enumerate(sorted(variables, key=lambda x: x["id"]), 1):
        print(f"  {i:>2}. {v['id']:<16} {short_label(v['label'], 36):<36} = {v['value']} {v['units']:<6}"
              f" conf {v['confidence']:.2f} {('dn:' + v['defined_name'][0]) if v['defined_name'] else ''}")

    print(f"\n C O E F F I C I E N T S\n  A. factor / property / limit cells   [{len(coeff_cells)}]\n")
    for i, v in enumerate(sorted(coeff_cells, key=lambda x: x["id"]), 1):
        print(f"  {i:>2}. {v['id']:<16} {short_label(v['label'], 36):<36} = {v['value']} {v['units']:<6}"
              f" conf {v['confidence']:.2f}")
    print(f"\n  B. numeric literals embedded inside formulas   [{len(embedded)}]\n")
    for lit in embedded:
        u = lit["used_in"][0]
        print(f"     {lit['literal']:<10} x{lit['occurrences']:<3} e.g. {u['cell']:<14} = {u['formula']}")

    print(f"\n C O N S T R A I N T S   (design checks / optimisation inequalities)   [{len(constraints)}]\n")
    for i, c in enumerate(constraints, 1):
        print(f"  {i:>2}. {c['id']:<16} {short_label(c['label'], 30):<30} {c['inequality']}")
        if c["g_form"]:
            print(f"       {c['g_form']}")

    interm = [c for c in cells.values() if c.kind == "intermediate"]
    amb = [c for c in cells.values() if c.kind == "ambiguous"]
    print(f"\n INTERMEDIATE / calculated cells: {len(interm)}   (see registry.json)")
    print(f" AMBIGUOUS (need your confirmation): {len(amb)}")
    for c in amb:
        print(f"     {c.id:<16} {short_label(c.label, 40)} value={c.value!r}")
    print(f"\n outputs -> {outdir}/  (registry.json, variables.json, coefficients.json, constraints.json)\n")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Identify variables, coefficients and constraints in an Excel bridge-design workbook")
    ap.add_argument("workbook", help="path to the .xlsx workbook")
    ap.add_argument("--confirm", action="store_true", help="ask you to confirm ambiguous cells in the terminal")
    ap.add_argument("--answers", help="JSON file with previously-supplied {'overrides': {...}, 'skip': [...]}")
    ap.add_argument("--outdir", default="classify_out", help="output directory for JSON files")
    ap.add_argument("--min-confidence", type=float, default=0.7, help="confidence below which a cell is ambiguous")
    args = ap.parse_args()

    ans = None
    if args.answers:
        with open(args.answers) as fh:
            ans = json.load(fh)
    classify_workbook(args.workbook, ans, args.min_confidence, args.outdir)
