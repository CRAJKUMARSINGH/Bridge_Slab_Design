#!/usr/bin/env python3
"""make_demo.py ΓÇö builds a small synthetic bridge-design workbook to exercise
classify_cells.py end-to-end.  All numbers are ILLUSTRATIVE, not a real design."""
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

wb = openpyxl.Workbook()

# ---------- Inputs sheet (design variables) ----------
ws_in = wb.active
ws_in.title = "Inputs"
ws_in["A1"] = "BRIDGE DESIGN - INPUT PARAMETERS"
labels = [
    ("Span length L (m)", 45.0),
    ("Deck width W (m)", 12.0),
    ("Girder spacing s (m)", 3.0),
    ("Number of girders n (-)", 4),
    ("Concrete strength fck (MPa)", 40.0),
    ("Steel grade fy (MPa)", 355.0),
    ("Deck thickness t (m)", 0.25),
    ("Live load UDL wl (kN/m)", 25.0),
    ("Live load concentrated Q (kN)", 150.0),
    ("Second moment of area I (mm4)", 6.5e10),
]
for i, (lbl, val) in enumerate(labels, start=3):
    ws_in.cell(row=i, column=1, value=lbl)
    ws_in.cell(row=i, column=2, value=val).number_format = "0.###"

# ---------- Factors sheet (coefficients) ----------
ws_f = wb.create_sheet("Factors")
ws_f["A1"] = "PARTIAL SAFETY & MATERIAL FACTORS"
factors = [
    ("Gamma_G (self weight)", 1.35),
    ("Gamma_Q (live load)", 1.5),
    ("Gamma_M (steel)", 1.1),
    ("Gamma_c (concrete)", 1.5),
    ("Elastic modulus E (MPa)", 210000.0),
    ("Unit weight concrete (kN/m3)", 25.0),
    ("Unit weight asphalt (kN/m3)", 24.0),
    ("Asphalt thickness (m)", 0.1),
    ("Deflection limit denominator (-)", 250.0),
]
for i, (lbl, val) in enumerate(factors, start=3):
    ws_f.cell(row=i, column=1, value=lbl)
    ws_f.cell(row=i, column=2, value=val)

# ---------- Calc sheet (formulas + checks) ----------
ws_c = wb.create_sheet("Calc")
ws_c["A1"] = "CALCULATIONS & CHECKS"

def L(r, lbl, formula, fmt="0.###"):
    ws_c.cell(row=r, column=1, value=lbl)
    ws_c.cell(row=r, column=2, value=formula).number_format = fmt

L(3, "Dead load slab g_slab (kN/m)", "=Inputs!B9*Factors!B7*Factors!B2")
L(4, "Dead load asphalt g_asph (kN/m)", "=Factors!B9*Factors!B8*Factors!B2")
L(5, "Design UDL incl. live w (kN/m)", "=B3+B4+Inputs!B10*Factors!B3")
L(6, "Max bending moment M (kN*m)", "=B5*Inputs!B3^2/8", "#,##0.0")
L(7, "Max shear V (kN)", "=B5*Inputs!B3/2+Inputs!B11*Factors!B3", "#,##0.0")
# constant, role deliberately ambiguous -> confirmation prompt
ws_c.cell(row=9, column=1, value="Chosen section modulus Wc (mm3)")
ws_c.cell(row=9, column=2, value=45000000).number_format = "#,##0"
# deliberately ambiguous constants (no keyword-suggestive labels)
ws_c.cell(row=13, column=1, value="Parameter alpha (-)")
ws_c.cell(row=13, column=2, value=0.9).number_format = "0.00"
ws_c.cell(row=21, column=1, value="Correction k3 (-)")
ws_c.cell(row=21, column=2, value=1.2).number_format = "0.00"
L(10, "Bending stress demand (MPa)", "=B6*1000000/B9", "0.0")
L(11, "Utilisation (bending) U_b (-)", "=B10/(Inputs!B8/Factors!B4)", "0.000")
L(12, "Bending check (-)", '=IF(B11<=1,"OK","NOT OK")')
L(14, "Deflection live load (mm)", "=5*Inputs!B10*(Inputs!B3*1000)^4/(384*Factors!B5*Inputs!B12)", "0.0")
L(15, "Deflection limit (mm)", "=Inputs!B3*1000/Factors!B10", "0.0")
L(16, "Deflection check (-)", '=IF(B14<=B15,"OK","NOT OK")')
L(18, "Shear stress demand (MPa)", "=B7*1000/(Inputs!B6*Inputs!B4*1000*Inputs!B9*1000)", "0.000")
# constant, code limit, deliberately ambiguous -> confirmation prompt
ws_c.cell(row=19, column=1, value="Shear capacity vc (MPa)")
ws_c.cell(row=19, column=2, value=2.5)
L(20, "Shear check (-)", '=IF(B18<=B19,"OK","NOT OK")')
L(22, "Governing utilisation (-)", "=MAX(B11,B18/B19)", "0.000")
L(23, "Adjusted governing utilisation (-)", "=B22*B13*B21", "0.000")

# defined names (typical for Excel model inputs)
wb.defined_names["SpanL"] = DefinedName("SpanL", attr_text="Inputs!$B$3")
wb.defined_names["steelgrade_fy"] = DefinedName("steelgrade_fy", attr_text="Inputs!$B$8")

wb.save("/home/user/demo_bridge_raw.xlsx")
print("saved /home/user/demo_bridge_raw.xlsx")
