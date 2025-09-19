"""
ENHANCED EXCEL GENERATOR
========================

Excel generation module inspired by BridgeSlabDesigner's advanced features
Integrates professional formatting, formulas, and comprehensive calculations
into the main Streamlit application.

Based on BridgeSlabDesigner/client/src/lib/excel-generator.ts
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, List, Optional
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
import math
from datetime import datetime
import io
import streamlit as st

@dataclass
class ExcelGenerationOptions:
    """Excel generation configuration options"""
    include_formulas: bool = True
    professional_formatting: bool = True
    auto_calculations: bool = True
    include_charts: bool = True
    detailed_estimates: bool = True

class EnhancedExcelGenerator:
    """
    Enhanced Excel generator with professional formatting
    Inspired by BridgeSlabDesigner's advanced capabilities
    """
    
    def __init__(self):
        self.workbook = Workbook()
        self.sheets_created = []
        
        # Professional color scheme
        self.colors = {
            'header_blue': 'FF4472C4',
            'header_green': 'FF70AD47', 
            'header_orange': 'FFED7D31',
            'header_red': 'FFC55A5A',
            'header_purple': 'FF5B9BD5',
            'light_gray': 'FFE7E6E6',
            'yellow_highlight': 'FFFFEB9C',
            'white': 'FFFFFFFF'
        }
        
        # Professional fonts
        self.fonts = {
            'header': Font(name='Arial', size=16, bold=True, color=self.colors['white']),
            'subheader': Font(name='Arial', size=12, bold=True),
            'normal': Font(name='Arial', size=10),
            'small': Font(name='Arial', size=9)
        }
    
    def generate_complete_bridge_excel(self, bridge_data: Dict[str, Any], 
                                     calculation_results: Dict[str, Any],
                                     options: ExcelGenerationOptions) -> bytes:
        """
        Generate complete bridge design Excel file with all sheets
        """
        
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        # Create all required sheets
        self._create_input_parameters_sheet(bridge_data, options)
        self._create_slab_bridge_design_sheet(bridge_data, calculation_results, options)
        self._create_pier_design_sheet(bridge_data, calculation_results, options)
        self._create_abutment_design_sheet(bridge_data, calculation_results, options)
        
        # Save to bytes buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_sheet_header(self, sheet, title: str, color_key: str, merge_range: str):
        """Create a professional sheet header"""
        sheet.merge_cells(merge_range)
        header_cell = sheet[merge_range.split(':')[0]]
        header_cell.value = title
        header_cell.font = self.fonts['header']
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.fill = PatternFill(start_color=self.colors[color_key], 
                                     end_color=self.colors[color_key], fill_type='solid')
    
    def _add_section_header(self, sheet, title: str, row: int) -> int:
        """Add a section header and return next row"""
        cell = sheet.cell(row=row, column=1)
        cell.value = title
        cell.font = self.fonts['subheader']
        cell.fill = PatternFill(start_color=self.colors['light_gray'], 
                              end_color=self.colors['light_gray'], fill_type='solid')
        return row + 1
    
    def _add_data_rows(self, sheet, data: List[List], start_row: int) -> int:
        """Add data rows and return next available row"""
        current_row = start_row
        for row_data in data:
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = self.fonts['normal']
            current_row += 1
        return current_row
    
    def _add_calculation_rows(self, sheet, calc_data: List[List], start_row: int, 
                            options: ExcelGenerationOptions) -> int:
        """Add calculation rows with optional formulas"""
        current_row = start_row
        for row_data in calc_data:
            # Unpack: [description, value, formula]
            description, value, formula = row_data[:3]
            
            # Description in column A
            desc_cell = sheet.cell(row=current_row, column=1)
            desc_cell.value = description
            desc_cell.font = self.fonts['normal']
            
            # Value in column B
            value_cell = sheet.cell(row=current_row, column=2)
            value_cell.value = value
            value_cell.font = self.fonts['normal']
            
            # Formula in column C (if enabled and provided)
            if options.include_formulas and formula:
                formula_cell = sheet.cell(row=current_row, column=3)
                if formula.startswith('='):
                    formula_cell.value = formula
                else:
                    formula_cell.value = f"={formula}"
                formula_cell.font = Font(name='Consolas', size=9)
            
            current_row += 1
        return current_row
    
    def _apply_borders(self, sheet, start_row: int, end_row: int, 
                      start_col: int, end_col: int):
        """Apply professional borders to a range"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
    
    def _create_input_parameters_sheet(self, bridge_data: Dict[str, Any], 
                                     options: ExcelGenerationOptions):
        """Create comprehensive input parameters sheet"""
        
        sheet = self.workbook.create_sheet('Input Parameters', 0)
        
        # Header
        self._create_sheet_header(sheet, 'BRIDGE DESIGN INPUT PARAMETERS', 
                                'header_blue', 'A1:F1')
        
        row = 3
        
        # Project Information Section
        row = self._add_section_header(sheet, 'PROJECT INFORMATION', row)
        
        project_info = [
            ['Project Name', bridge_data.get('bridge_name', 'Unnamed Project')],
            ['Location', bridge_data.get('location', 'Not Specified')],
            ['Engineer', 'Bridge Design Application'],
            ['Date', datetime.now().strftime('%d-%m-%Y')],
            ['Drawing No.', f"BD-{datetime.now().strftime('%Y%m%d')}-001"],
            ['Revision', 'Rev-0']
        ]
        
        row = self._add_data_rows(sheet, project_info, row)
        
        # Bridge Configuration Section
        row += 1
        row = self._add_section_header(sheet, 'BRIDGE CONFIGURATION', row)
        
        bridge_config = [
            ['Number of Spans', bridge_data.get('num_spans', 3)],
            ['Effective Span (m)', bridge_data.get('effective_span', 9.6)],
            ['Bridge Width (m)', bridge_data.get('bridge_width', 12.0)],
            ['Pier Spacing C/C (m)', bridge_data.get('pier_spacing_cc', 11.1)],
            ['Pier Cap Width (m)', bridge_data.get('pier_cap_width', 15.0)],
            ['Skew Angle (°)', bridge_data.get('skew_angle', 0.0)]
        ]
        
        row = self._add_data_rows(sheet, bridge_config, row)
        
        # Hydraulic Parameters Section
        row += 1
        row = self._add_section_header(sheet, 'HYDRAULIC PARAMETERS', row)
        
        hydraulic_params = [
            ['Design Discharge (Cumecs)', bridge_data.get('discharge', 1265.76)],
            ['Design Velocity (m/s)', bridge_data.get('design_velocity', 3.5)],
            ['HFL (m)', bridge_data.get('hfl', 101.2)],
            ['Manning\'s n', bridge_data.get('manning_n', 0.033)],
            ['Afflux (m)', bridge_data.get('afflux', 0.083)]
        ]
        
        row = self._add_data_rows(sheet, hydraulic_params, row)
        
        # Soil Parameters Section
        row += 1
        row = self._add_section_header(sheet, 'SOIL PARAMETERS', row)
        
        soil_params = [
            ['Safe Bearing Capacity (kN/m²)', bridge_data.get('safe_bearing_capacity', 450.0)],
            ['Angle of Friction (°)', bridge_data.get('angle_of_friction', 30.0)],
            ['Unit Weight (kN/m³)', bridge_data.get('unit_weight', 18.0)],
            ['Coefficient of Friction', bridge_data.get('coefficient_of_friction', 0.6)]
        ]
        
        row = self._add_data_rows(sheet, soil_params, row)
        
        # Material Properties Section
        row += 1
        row = self._add_section_header(sheet, 'MATERIAL PROPERTIES', row)
        
        material_props = [
            ['Concrete Grade', bridge_data.get('concrete_grade', 'M25')],
            ['Steel Grade', bridge_data.get('steel_grade', 'Fe415')],
            ['Density of Concrete (kN/m³)', 25.0],
            ['Density of Steel (kN/m³)', 78.5]
        ]
        
        row = self._add_data_rows(sheet, material_props, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 15
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 4)
    
    def _create_slab_bridge_design_sheet(self, bridge_data: Dict[str, Any], 
                                       calculation_results: Dict[str, Any],
                                       options: ExcelGenerationOptions):
        """Create detailed slab bridge design calculations sheet"""
        
        sheet = self.workbook.create_sheet('Slab Bridge Design')
        
        self._create_sheet_header(sheet, 'SLAB BRIDGE DESIGN CALCULATIONS', 
                                'header_green', 'A1:G1')
        
        row = 3
        
        # Design Parameters
        row = self._add_section_header(sheet, 'DESIGN PARAMETERS', row)
        
        effective_span = bridge_data.get('effective_span', 9.6)
        bridge_width = bridge_data.get('bridge_width', 12.0)
        slab_thickness = 0.75  # meters
        
        design_params = [
            ['Effective Span (m)', effective_span],
            ['Bridge Width (m)', bridge_width],
            ['Slab Thickness (mm)', slab_thickness * 1000],
            ['Total Bridge Length (m)', effective_span * bridge_data.get('num_spans', 3)],
            ['Deck Area (m²)', bridge_width * effective_span * bridge_data.get('num_spans', 3)]
        ]
        
        row = self._add_data_rows(sheet, design_params, row)
        
        # Load Calculations with Formulas
        row += 1
        row = self._add_section_header(sheet, 'LOAD CALCULATIONS', row)
        
        # Dead Load calculations
        dead_load_slab = slab_thickness * 25  # kN/m²
        wearing_coat = 0.065 * 22  # 65mm @ 22 kN/m³
        total_dl = dead_load_slab + wearing_coat
        live_load = 15  # IRC Class AA
        total_load = total_dl + live_load
        
        load_calcs = [
            ['Self Weight of Slab (kN/m²)', dead_load_slab, f'={slab_thickness}*25'],
            ['Wearing Coat (kN/m²)', wearing_coat, '=0.065*22'],
            ['Total Dead Load (kN/m²)', total_dl, f'=B{row+1}+B{row+2}' if options.include_formulas else ''],
            ['Live Load IRC Class AA (kN/m²)', live_load, '=15'],
            ['Total Load (kN/m²)', total_load, f'=B{row+3}+B{row+4}' if options.include_formulas else '']
        ]
        
        row = self._add_calculation_rows(sheet, load_calcs, row, options)
        
        # Moment and Shear Calculations
        row += 1
        row = self._add_section_header(sheet, 'MOMENT AND SHEAR CALCULATIONS', row)
        
        # Structural analysis
        max_moment = (total_load * effective_span**2) / 8  # Simply supported
        max_shear = (total_load * effective_span) / 2
        
        moment_shear_calcs = [
            ['Maximum Bending Moment (kN.m/m)', max_moment, f'=B{row-1}*{effective_span}^2/8'],
            ['Maximum Shear Force (kN/m)', max_shear, f'=B{row-1}*{effective_span}/2'],
            ['Design Moment (kN.m/m)', max_moment * 1.5, f'=B{row+1}*1.5'],
            ['Design Shear (kN/m)', max_shear * 1.5, f'=B{row+2}*1.5']
        ]
        
        row = self._add_calculation_rows(sheet, moment_shear_calcs, row, options)
        
        # Steel Design Calculations
        row += 1
        row = self._add_section_header(sheet, 'STEEL DESIGN CALCULATIONS', row)
        
        fck = 25  # N/mm²
        fy = 415   # N/mm²
        effective_depth = 675  # mm (750 - 75 cover)
        
        # Required steel area calculation
        design_moment_nmm = max_moment * 1.5 * 1e6  # Convert to N.mm
        ast_required = design_moment_nmm / (0.87 * fy * effective_depth * 0.9)
        ast_min = 0.12 * bridge_width * 1000 * slab_thickness * 1000 / 100  # 0.12% of gross area
        ast_provided = max(ast_required, ast_min)
        
        steel_calcs = [
            ['Effective Depth (mm)', effective_depth],
            ['fck (N/mm²)', fck],
            ['fy (N/mm²)', fy],
            ['Required Steel Area (mm²/m)', ast_required],
            ['Minimum Steel Area (mm²/m)', ast_min],
            ['Provided Steel Area (mm²/m)', ast_provided],
        ]
        
        row = self._add_data_rows(sheet, steel_calcs, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 40
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 25
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 7)
    
    def _create_hydraulic_design_sheet(self, bridge_data: Dict[str, Any], 
                                     calculation_results: Dict[str, Any],
                                     options: ExcelGenerationOptions):
        """Create hydraulic design calculations sheet with Excel formulas"""
        
        sheet = self.workbook.create_sheet('Hydraulic Design')
        
        self._create_sheet_header(sheet, 'HYDRAULIC DESIGN CALCULATIONS', 
                                'header_blue', 'A1:H1')
        
        row = 3
        
        # Bridge Parameters
        row = self._add_section_header(sheet, 'HYDRAULIC PARAMETERS', row)
        
        discharge = bridge_data.get('discharge', 1265.76)
        velocity = bridge_data.get('design_velocity', 3.5)
        manning_n = bridge_data.get('manning_n', 0.033)
        slope = 1/960  # Typical slope
        
        # Calculate hydraulic properties
        area = discharge / velocity
        perimeter = 2 * math.sqrt(area * 4)  # Approximate rectangular section
        hydraulic_radius = area / perimeter
        velocity_manning = (1/manning_n) * (hydraulic_radius**(2/3)) * (slope**0.5)
        regime_width = 4.8 * math.sqrt(discharge)
        
        hydraulic_calcs = [
            ['S.No.', 'Description', 'Formula', 'Calculation', 'Result', 'Unit', 'Remarks'],
            ['1', 'Design Discharge', 'Given', f'Q = {discharge}', discharge, 'Cumecs', 'From Hydrology'],
            ['2', 'Design Velocity', 'Given', f'V = {velocity}', velocity, 'm/s', 'From Flow Analysis'],
            ['3', 'Cross-sectional Area', 'Q/V', f'A = {discharge}/{velocity}', area, 'm²', 'Continuity Equation'],
            ['4', 'Wetted Perimeter', 'Survey', f'P = {perimeter:.2f}', perimeter, 'm', 'From Survey Data'],
            ['5', 'Hydraulic Radius', 'A/P', f'R = {area:.2f}/{perimeter:.2f}', hydraulic_radius, 'm', 'R = A/P'],
            ['6', "Manning's Velocity", '(1/n)*R^(2/3)*S^(1/2)', f'V = (1/{manning_n})*{hydraulic_radius:.3f}^(2/3)*{slope:.6f}^(1/2)', velocity_manning, 'm/s', "Manning's Formula"],
            ['7', 'Regime Width (Lacey)', '4.8*Q^0.5', f'Wr = 4.8*{discharge}^0.5', regime_width, 'm', "Lacey's Formula"]
        ]
        
        # Add header row
        for col_idx, header in enumerate(hydraulic_calcs[0], 1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['light_gray'], 
                                  end_color=self.colors['light_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Add calculation rows
        for calc_row in hydraulic_calcs[1:]:
            for col_idx, value in enumerate(calc_row, 1):
                cell = sheet.cell(row=row, column=col_idx)
                if isinstance(value, (int, float)) and col_idx == 5:  # Result column
                    cell.value = f"{value:.3f}" if isinstance(value, float) else value
                else:
                    cell.value = value
                cell.font = self.fonts['normal']
                
                # Center align S.No. column
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal='center')
            row += 1
        
        # Set column widths
        column_widths = [8, 25, 20, 30, 15, 10, 20]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = width
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row-1, 1, 7)
    
    def _create_stability_analysis_sheet(self, bridge_data: Dict[str, Any], 
                                       calculation_results: Dict[str, Any],
                                       options: ExcelGenerationOptions):
        """Create stability analysis sheet"""
        
        sheet = self.workbook.create_sheet('Stability Analysis')
        
        self._create_sheet_header(sheet, 'STABILITY ANALYSIS CALCULATIONS', 
                                'header_red', 'A1:G1')
        
        row = 3
        
        # Overturning Analysis
        row = self._add_section_header(sheet, 'OVERTURNING STABILITY ANALYSIS', row)
        
        # Sample loads for demonstration
        vertical_load = 8500.0  # kN
        horizontal_load = 850.0  # kN
        height = 6.0  # m
        foundation_width = 3.0  # m
        
        moment_overturning = horizontal_load * height
        moment_restoring = vertical_load * (foundation_width / 2)
        fs_overturning = moment_restoring / moment_overturning
        
        overturning_calcs = [
            ['Overturning Moment (kN.m)', moment_overturning, f'={horizontal_load}*{height}'],
            ['Restoring Moment (kN.m)', moment_restoring, f'={vertical_load}*{foundation_width}/2'],
            ['Factor of Safety against Overturning', fs_overturning, f'=B{row+2}/B{row+1}'],
            ['Status', 'SAFE' if fs_overturning >= 2.0 else 'UNSAFE', f'=IF(B{row+3}>=2,"SAFE","UNSAFE")']
        ]
        
        row = self._add_calculation_rows(sheet, overturning_calcs, row, options)
        
        # Sliding Analysis
        row += 2
        row = self._add_section_header(sheet, 'SLIDING STABILITY ANALYSIS', row)
        
        friction_angle = bridge_data.get('angle_of_friction', 30.0)
        coefficient_friction = math.tan(math.radians(friction_angle))
        resisting_force = coefficient_friction * vertical_load
        fs_sliding = resisting_force / horizontal_load
        
        sliding_calcs = [
            ['Coefficient of Friction', coefficient_friction, f'=TAN(RADIANS({friction_angle}))'],
            ['Resisting Force (kN)', resisting_force, f'=B{row+1}*{vertical_load}'],
            ['Factor of Safety against Sliding', fs_sliding, f'=B{row+2}/{horizontal_load}'],
            ['Status', 'SAFE' if fs_sliding >= 1.5 else 'UNSAFE', f'=IF(B{row+3}>=1.5,"SAFE","UNSAFE")']
        ]
        
        row = self._add_calculation_rows(sheet, sliding_calcs, row, options)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 25
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 7)
    
    def _create_steel_design_sheet(self, bridge_data: Dict[str, Any], 
                                 calculation_results: Dict[str, Any],
                                 options: ExcelGenerationOptions):
        """Create steel design calculations sheet"""
        
        sheet = self.workbook.create_sheet('Steel Design')
        
        self._create_sheet_header(sheet, 'STEEL DESIGN CALCULATIONS', 
                                'header_orange', 'A1:F1')
        
        row = 3
        
        # Material Properties
        row = self._add_section_header(sheet, 'MATERIAL PROPERTIES', row)
        
        concrete_grade = bridge_data.get('concrete_grade', 'M25')
        steel_grade = bridge_data.get('steel_grade', 'Fe415')
        fck = 25 if concrete_grade == 'M25' else 30
        fy = 415 if steel_grade == 'Fe415' else 500
        
        material_props = [
            ['Concrete Grade', concrete_grade],
            ['fck (N/mm²)', fck],
            ['Steel Grade', steel_grade],
            ['fy (N/mm²)', fy],
            ['Factor of Safety (concrete)', 1.5],
            ['Factor of Safety (steel)', 1.15]
        ]
        
        row = self._add_data_rows(sheet, material_props, row)
        
        # Reinforcement Design
        row += 1
        row = self._add_section_header(sheet, 'REINFORCEMENT DESIGN', row)
        
        # Sample design moment
        design_moment = 150.0  # kN.m
        effective_depth = 675  # mm
        width = 1000  # mm (per meter width)
        
        # Steel area calculation
        design_moment_nmm = design_moment * 1e6
        ast_required = design_moment_nmm / (0.87 * fy * effective_depth * 0.9)
        ast_min = 0.12 * width * effective_depth / 100
        ast_provided = max(ast_required, ast_min)
        
        # Bar selection
        bar_diameter = 16  # mm
        bar_area = math.pi * (bar_diameter/2)**2
        number_of_bars = math.ceil(ast_provided / bar_area)
        spacing = 1000 / number_of_bars
        
        reinforcement_design = [
            ['Design Moment (kN.m)', design_moment],
            ['Effective Depth (mm)', effective_depth],
            ['Width (mm)', width],
            ['Required Steel Area (mm²)', ast_required],
            ['Minimum Steel Area (mm²)', ast_min],
            ['Provided Steel Area (mm²)', ast_provided],
            ['Bar Diameter (mm)', bar_diameter],
            ['Number of Bars', number_of_bars],
            ['Spacing (mm)', spacing],
        ]
        
        row = self._add_data_rows(sheet, reinforcement_design, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 6)
    
    def _create_pier_design_sheet(self, bridge_data: Dict[str, Any], 
                                calculation_results: Dict[str, Any],
                                options: ExcelGenerationOptions):
        """Create pier design sheet"""
        
        sheet = self.workbook.create_sheet('Pier Design')
        
        self._create_sheet_header(sheet, 'PIER DESIGN CALCULATIONS', 
                                'header_orange', 'A1:F1')
        
        row = 3
        
        # Pier Geometry
        row = self._add_section_header(sheet, 'PIER GEOMETRY', row)
        
        num_piers = bridge_data.get('num_spans', 3) - 1
        pier_height = bridge_data.get('hfl', 101.2) - bridge_data.get('bedLevel', 294.0) + 2.0
        pier_width = 2.0  # m
        pier_thickness = 1.5  # m
        
        pier_geometry = [
            ['Number of Piers', num_piers],
            ['Pier Height (m)', pier_height],
            ['Pier Width (m)', pier_width],
            ['Pier Thickness (m)', pier_thickness],
            ['Volume per Pier (m³)', pier_width * pier_thickness * pier_height],
            ['Total Pier Volume (m³)', num_piers * pier_width * pier_thickness * pier_height]
        ]
        
        row = self._add_data_rows(sheet, pier_geometry, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 6)
    
    def _create_abutment_design_sheet(self, bridge_data: Dict[str, Any], 
                                    calculation_results: Dict[str, Any],
                                    options: ExcelGenerationOptions):
        """Create abutment design sheet"""
        
        sheet = self.workbook.create_sheet('Abutment Design')
        
        self._create_sheet_header(sheet, 'ABUTMENT DESIGN CALCULATIONS', 
                                'header_purple', 'A1:F1')
        
        row = 3
        
        # Abutment Geometry
        row = self._add_section_header(sheet, 'ABUTMENT GEOMETRY', row)
        
        num_abutments = 2
        abutment_height = bridge_data.get('hfl', 101.2) - bridge_data.get('bedLevel', 294.0) + 2.0
        abutment_length = 3.0  # m
        abutment_thickness = 2.0  # m
        
        abutment_geometry = [
            ['Number of Abutments', num_abutments],
            ['Abutment Height (m)', abutment_height],
            ['Abutment Length (m)', abutment_length],
            ['Abutment Thickness (m)', abutment_thickness],
            ['Volume per Abutment (m³)', abutment_length * abutment_thickness * abutment_height],
            ['Total Abutment Volume (m³)', num_abutments * abutment_length * abutment_thickness * abutment_height]
        ]
        
        row = self._add_data_rows(sheet, abutment_geometry, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 6)
    
    def _create_foundation_design_sheet(self, bridge_data: Dict[str, Any], 
                                      calculation_results: Dict[str, Any],
                                      options: ExcelGenerationOptions):
        """Create foundation design sheet"""
        
        sheet = self.workbook.create_sheet('Foundation Design')
        
        self._create_sheet_header(sheet, 'FOUNDATION DESIGN CALCULATIONS', 
                                'header_red', 'A1:F1')
        
        row = 3
        
        # Foundation Parameters
        row = self._add_section_header(sheet, 'FOUNDATION PARAMETERS', row)
        
        foundation_type = 'Spread Footings'
        foundation_depth = 3.0  # m
        foundation_length = 3.0  # m
        foundation_width = 2.5  # m
        sbc = bridge_data.get('safe_bearing_capacity', 450.0)
        
        foundation_data = [
            ['Foundation Type', foundation_type],
            ['Foundation Depth (m)', foundation_depth],
            ['Foundation Length (m)', foundation_length], 
            ['Foundation Width (m)', foundation_width],
            ['Safe Bearing Capacity (kN/m²)', sbc],
            ['Volume per Foundation (m³)', foundation_length * foundation_width * foundation_depth]
        ]
        
        row = self._add_data_rows(sheet, foundation_data, row)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 20
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row, 1, 6)
    
    def _create_general_abstract_sheet(self, bridge_data: Dict[str, Any], 
                                     calculation_results: Dict[str, Any],
                                     options: ExcelGenerationOptions):
        """Create general abstract cost sheet"""
        
        sheet = self.workbook.create_sheet('General Abstract')
        
        self._create_sheet_header(sheet, 'GENERAL ABSTRACT OF COST', 
                                'header_blue', 'A1:D1')
        
        row = 3
        
        # Headers
        headers = ['S.No.', 'Item Description', 'Amount (₹)', 'Percentage (%)']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['light_gray'], 
                                  end_color=self.colors['light_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Cost breakdown
        structure_cost = 5000000  # ₹50 Lakh
        foundation_cost = 3000000  # ₹30 Lakh
        total_cost = structure_cost + foundation_cost
        
        abstract_items = [
            ['1', 'Sub-structure', foundation_cost, f'{(foundation_cost/total_cost*100):.1f}'],
            ['2', 'Super-structure', structure_cost, f'{(structure_cost/total_cost*100):.1f}'],
            ['', 'TOTAL', total_cost, '100.0']
        ]
        
        for item_data in abstract_items:
            for col_idx, value in enumerate(item_data, 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = value
                if row == len(abstract_items) + 3:  # Total row
                    cell.font = self.fonts['subheader']
                    cell.fill = PatternFill(start_color=self.colors['yellow_highlight'], 
                                          end_color=self.colors['yellow_highlight'], fill_type='solid')
                else:
                    cell.font = self.fonts['normal']
            row += 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 15
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row-1, 1, 4)
    
    def _create_detailed_estimate_sheet(self, bridge_data: Dict[str, Any], 
                                      calculation_results: Dict[str, Any],
                                      options: ExcelGenerationOptions):
        """Create detailed estimate sheet"""
        
        sheet = self.workbook.create_sheet('Detailed Estimate')
        
        self._create_sheet_header(sheet, 'DETAILED ESTIMATE', 
                                'header_green', 'A1:F1')
        
        row = 3
        
        # Headers
        headers = ['S.No.', 'Description of Item', 'Unit', 'Quantity', 'Rate (₹)', 'Amount (₹)']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['light_gray'], 
                                  end_color=self.colors['light_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Sample detailed items
        detailed_items = [
            ['1', 'Excavation for Foundation', 'm³', '150', '150', '22500'],
            ['2', 'Concrete in Foundation (M25)', 'm³', '45', '4500', '202500'],
            ['3', 'Concrete in Pier (M25)', 'm³', '60', '4500', '270000'],
            ['4', 'Concrete in Deck Slab (M25)', 'm³', '85', '4500', '382500'],
            ['5', 'Steel Reinforcement (Fe415)', 'tonne', '12', '65000', '780000'],
            ['6', 'Formwork', 'm²', '450', '350', '157500']
        ]
        
        for item_data in detailed_items:
            for col_idx, value in enumerate(item_data, 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = self.fonts['normal']
                if col_idx == 1:  # S.No.
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx >= 4:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 10
        sheet.column_dimensions['D'].width = 15
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 20
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row-1, 1, 6)
    
    def _create_quantity_measurements_sheet(self, bridge_data: Dict[str, Any], 
                                          calculation_results: Dict[str, Any],
                                          options: ExcelGenerationOptions):
        """Create quantity measurements sheet"""
        
        sheet = self.workbook.create_sheet('Quantity Measurements')
        
        self._create_sheet_header(sheet, 'QUANTITY MEASUREMENTS', 
                                'header_orange', 'A1:G1')
        
        row = 3
        
        # Headers
        headers = ['S.No.', 'Item', 'Length', 'Width', 'Height', 'Nos.', 'Total Qty']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = self.fonts['subheader']
            cell.fill = PatternFill(start_color=self.colors['light_gray'], 
                                  end_color=self.colors['light_gray'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
        
        # Sample measurements
        effective_span = bridge_data.get('effective_span', 9.6)
        bridge_width = bridge_data.get('bridge_width', 12.0)
        num_spans = bridge_data.get('num_spans', 3)
        
        measurements = [
            ['1', 'Slab Concrete', f'{effective_span * num_spans:.1f}', f'{bridge_width:.1f}', '0.750', '1', f'{effective_span * num_spans * bridge_width * 0.75:.2f}'],
            ['2', 'Pier Concrete', '2.0', '1.5', '6.0', f'{num_spans-1}', f'{(num_spans-1) * 2.0 * 1.5 * 6.0:.2f}'],
            ['3', 'Foundation Excavation', '3.0', '2.5', '3.0', f'{num_spans-1}', f'{(num_spans-1) * 3.0 * 2.5 * 3.0:.2f}']
        ]
        
        for measurement in measurements:
            for col_idx, value in enumerate(measurement, 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.value = value
                cell.font = self.fonts['normal']
                if col_idx == 1 or col_idx == 6:  # S.No. and Nos.
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx >= 3:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Set column widths
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 8
        sheet.column_dimensions['G'].width = 15
        
        if options.professional_formatting:
            self._apply_borders(sheet, 1, row-1, 1, 7)